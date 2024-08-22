from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import OuterRef, Subquery, F
from last_kgo.settings import *
from django.core.paginator import Paginator
from .models import *
from .forms import *
import openpyxl


def home(request):
    return render(request, 'base.html')


def export_to_excel(data_list):
    # Create an in-memory output file for the new workbook.
    output = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    output['Content-Disposition'] = 'attachment; filename=universities_data.xlsx'

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Universities Data"

    # Define the titles for columns
    columns = [
        'Program Code', 'Department Name', 'University Name',
        'University Type', 'Burs Türü', 'Genel Kontenjan', 'İlk Yerleşme Oranı', 'Puan Türü', 'Yıl'
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Write data to cells
    for item in data_list:
        row_num += 1
        row = [
            item.get('pro_code'), item.get('bolum_adi'), item.get('üniversite'),
            item.get('üniversite_türü'), item.get('burs_türü'), item.get('genel_kontenjan'),
            item.get('i_lk_yerleşme_oranı'),
            item.get('puan_türü'), item.get('yil')
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Save the workbook to the response
    wb.save(output)
    return output


def filters_page(request):
    form = GeneralInformationForm(request.GET)
    data_list = GeneralInformation.objects.all().order_by('üniversite')

    if form.is_valid():
        university_type = form.cleaned_data.get('university_type')
        university_name = form.cleaned_data.get('university_name')
        year = form.cleaned_data.get('year')
        scholarship = form.cleaned_data.get('scholarship')
        major_name = form.cleaned_data.get('major_name')

        if university_type:
            data_list = data_list.filter(üniversite_türü__in=university_type)
        if university_name:
            data_list = data_list.filter(üniversite__in=university_name)
        if major_name:
            data_list = data_list.filter(bolum_adi__in=major_name)
        if year:
            try:
                data_list = data_list.filter(yil__in=year)
            except ValueError:
                print("passing..")
                pass
        if scholarship:
            data_list = data_list.filter(burs_türü__in=scholarship)

        request.session['filtered_data'] = []
        request.session['filtered_data'] = list(data_list.values())

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        form.fields['university_type'].choices = [('', 'Hepsi')] + [(choice, choice) for choice in
                                                                    data_list.values_list('üniversite_türü',
                                                                                          flat=True).distinct()]
        form.fields['university_name'].choices = [('', 'Hepsi')] + [(choice, choice) for choice in
                                                                    data_list.values_list('üniversite',
                                                                                          flat=True).distinct()]
        form.fields['major_name'].choices = [('', 'Hepsi')] + [(choice, choice) for choice in
                                                                    data_list.values_list('bolum_adi',
                                                                                          flat=True).distinct()]

        form.fields['year'].choices = [('', 'Hepsi')] + [(choice, choice) for choice in
                                                         data_list.values_list('yil', flat=True).distinct()]
        form.fields['scholarship'].choices = [('', 'Hepsi')] + [(choice, choice) for choice in
                                                                data_list.values_list('burs_türü',
                                                                                      flat=True).distinct()]

        return render(request, 'filters.html', {'form': form})

    if 'export' in request.GET:
        filtered_data = request.session.get('filtered_data', [])
        print("filtered", filtered_data)
        return export_to_excel(filtered_data)

    paginator = Paginator(data_list, 50)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'form': form
    }

    return render(request, 'filters.html', context)


def contact_us(request):
    return render(request, 'contact.html')


def old_hc_data_page(request):
    return render(request, 'hc.html')


def high_school_page(request):
    high_school_data = HighSchoolTable.objects.all().order_by('university')
    paginator = Paginator(high_school_data, 50)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    return render(request, 'high_school.html', {'high_school_data': page_obj})


def preference_tendency_view(request):
    form = PreferenceTendencyFilterForm(request.GET)
    data_list = VwPreferenceTendencySamePrograms.objects.all()

    if form.is_valid():
        university = form.cleaned_data.get('university')
        major_name = form.cleaned_data.get('major_name')
        year = form.cleaned_data.get('year')

        if university:
            data_list = data_list.filter(üniversite__in=university)
        if major_name:
            data_list = data_list.filter(bolum_adi__in=major_name)
        if year:
            try:
                data_list = data_list.filter(yil__in=year)
            except ValueError:
                print("passing..")
                pass

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        form.fields['university'].choices = [('', 'Hepsi')] + [(choice, choice) for choice in
                                                                    data_list.values_list('üniversite',
                                                                                          flat=True).distinct()]
        form.fields['major_name'].choices = [('', 'Hepsi')] + [(choice, choice) for choice in
                                                                    data_list.values_list('bolum_adi',
                                                                                          flat=True).distinct()]

        form.fields['year'].choices = [('', 'Hepsi')] + [(choice, choice) for choice in
                                                         data_list.values_list('yil', flat=True).distinct()]
        return render(request, 'preference_tendancy.html', {'form': form})

    paginator = Paginator(data_list, 50)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'form': form,
        'page_obj': page_obj,
    }

    return render(request, 'preference_tendancy.html', context)


def tercih_sihirbazi(request):
    tercih_sihirbazi_data = TercihSihirbazi.objects.all().order_by('universite')

    paginator = Paginator(tercih_sihirbazi_data, 50)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
    }

    return render(request, 'tercih_sihirbazi.html', context)
